from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule

from Processing.charts import create_chart

#Colors
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
kpi_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

def style_header(ws, row = 1):
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

def auto_adjust_col_width(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter

        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        ws.column_dimensions[col_letter].width = max_length + 3

def generate_report(df, summary):
    wb = Workbook()

    #Sheet 1 : Raw Data
    ws1 = wb.active
    ws1.title = "Raw Data"

    for r in dataframe_to_rows(df, index=False, header=True):
        ws1.append(r)
    
    style_header(ws1)
    ws1.freeze_panes="A2"
    ws1.auto_filter.ref = ws1.dimensions
    

    #Conditional Formatting (Rating Colors)
    for col_idx, col_name in enumerate(df.columns, start=1):
        if col_name == "Rating":
            col_letter = ws1.cell(row=1, column=col_idx).column_letter
    
            ws1.conditional_formatting.add(
              f"{col_letter}2:{col_letter}{ws1.max_row}",
              CellIsRule(operator='greaterThanOrEqual', formula=['4'], fill=PatternFill(start_color="C6EFCE", fill_type="solid"))
            )
            ws1.conditional_formatting.add(
              f"{col_letter}2:{col_letter}{ws1.max_row}",
              CellIsRule(operator='lessThan', formula=['4'], fill=PatternFill(start_color="FFC7CE", fill_type="solid"))
            )
    auto_adjust_col_width(ws1)

    #Sheet 3: City Analysis

    ws2 = wb.create_sheet("City Analysis")

    ws2.append(["City","Total Leads","Average Rating"])
    style_header(ws2)

    for city, row in summary["city_analysis"].iterrows():
        ws2.append([
            city,
            row["Total Leads"],
            round(row["Rating"],2) if row["Rating"] else 0
        ])

    auto_adjust_col_width(ws2)

    #Create Charts
    charts = create_chart(ws2)

 #Sheet 3: Dashboard
    ws3 = wb.create_sheet("Dashboard")

    ws3["A1"]="Lead Generation Dashboard"
    ws3["A1"].font= Font(size = 16, bold = True)

    ws3["A3"]="Business Type"
    ws3["B3"]=summary["business_type"]

    ws3["A4"]="Total Leads"
    ws3["B4"]=summary["total_leads"]

    ws3["A5"]="Cities"
    ws3["B5"]=summary["cities"]

    ws3["A6"]="Average Ratings"
    ws3["B6"]=round(summary["avg_rating"],2)

    ws3["A7"]="Website"
    ws3["B7"]=summary["website_count"]

    ws3["A8"]="Email"
    ws3["B8"]=summary["email_count"]

    ws3.add_chart(charts["leads_chart"],"D3")
    ws3.add_chart(charts["rating_chart"],"D18")
    ws3.add_chart(charts["pie_chart"],"K3")

    auto_adjust_col_width(ws3)

    #Sheet 4: Top Companies
    ws4 = wb.create_sheet("Top Companies")

    ws4.append(["Company", "City", "Rating", "Website","Email"])
    style_header(ws4)

    for _, row in summary["top_companies"].iterrows():
        ws4.append([
            row["Company"],
            row["City"],
            row["Rating"],
            row["Website"],
            row["Email"]
        ])

    ws4.auto_filter.ref = ws4.dimensions 
    auto_adjust_col_width(ws4)

    wb.save("C:/Users/hp/OneDrive/Desktop/Free/Lead_Generation_Automation/Data/report.xlsx")

    print("Report Generated Successfully!")