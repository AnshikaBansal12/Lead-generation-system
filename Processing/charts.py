from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.layout import Layout, ManualLayout

def create_chart(ws_city):
    charts={}

    #Chart1- leads per city
    chart1 = BarChart()
    chart1.title = "Leads per City"

    #Axis Title
    chart1.x_axis.title = "City"
    chart1.y_axis.title = "Number of Leads"

    #Axis Visibility
    chart1.x_axis.delete = False
    chart1.y_axis.delete = False

    #Tick marks
    chart1.x_axis.majorTickMark = "out"
    chart1.y_axis.majorTickMark = "out"

    #Label Position
    chart1.x_axis.tickLblPos = "low"
    chart1.y_axis.tickLblPos = "nextTo"

    #Chart Size
    chart1.width = 10
    chart1.height = 7

    chart1.style = 10

    chart1.layout = Layout(
        manualLayout=ManualLayout(
            x=0.04,
            y=0.05,
            w=0.8,
            h=0.6
        )
    )
    #data

    data1 = Reference(ws_city, min_col=2, min_row=1, max_row=ws_city.max_row)
    cats1 = Reference(ws_city, min_col=1, min_row=2, max_row=ws_city.max_row)

    chart1.add_data(data1, titles_from_data=True)
    chart1.set_categories(cats1)

    #Show values on bars
    chart1.dLbls = DataLabelList()
    chart1.dLbls.showVal = True

    chart1.y_axis.majorGridlines = None
    chart1.gapWidth = 150

    #print(ws_city.max_row)
    #print(ws_city.cell(row=2, column=1).value)

    charts["leads_chart"] = chart1

    #Chart2: Average Rating
    chart2 = BarChart()
    chart2.title = "Average Rating by City"

    #Axis Title
    chart2.x_axis.title = "City"
    chart2.y_axis.title = "Rating"

    #Axis Visisbility
    chart2.x_axis.delete = False
    chart2.y_axis.delete = False

    #Axis Marks
    chart2.x_axis.majorTickMark = "out"
    chart2.y_axis.majorTickMark = "out"

    #Label Position
    chart2.x_axis.tickLblPos = "low"
    chart2.y_axis.tickLblPos = "nextTo"

    #chart size
    chart2.width = 12
    chart2.height = 8

    chart2.style = 10

    #Chart layout
    chart2.layout = Layout(
        manualLayout=ManualLayout(
            x=0.03,
            y=0.05,
            w=0.8,
            h=0.7
        )
    )

    #data
    data2 = Reference(ws_city, min_col=3, min_row=1, max_row=ws_city.max_row)
    cats2 = Reference(ws_city, min_col=1, min_row=2, max_row=ws_city.max_row)

    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats2)

    #show values on bar
    chart2.dLbls = DataLabelList()
    chart2.dLbls.showVal = True

    chart2.y_axis.majorGridlines = None
    chart2.gapWidth = 150

    charts["rating_chart"] = chart2

    #Chart3: Pie Chart
    chart3 = PieChart()
    chart3.title = "Lead Distribution By City"

    #chart size
    chart3.width = 10
    chart3.height = 7

    chart3.style = 10
    
    #Chart layout
    chart3.layout = Layout(
        manualLayout=ManualLayout(
            x=0.03,
            y=0.05,
            w=0.8,
            h=0.7
        )
    )

    data3 = Reference(ws_city, min_col=2,min_row=1,max_row=ws_city.max_row)
    cats3 = Reference(ws_city, min_col=1, min_row=2, max_row=ws_city.max_row)

    chart3.add_data(data3, titles_from_data=True)
    chart3.set_categories(cats3)

    chart3.dLbls = DataLabelList()
    chart3.dLbls.showPercent = True

    charts["pie_chart"] = chart3

    return charts